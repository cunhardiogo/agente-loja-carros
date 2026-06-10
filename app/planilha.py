import hashlib
import io
import logging
import re
import unicodedata
from datetime import date, datetime, time

import httpx
import openpyxl

from . import db
from .config import settings

log = logging.getLogger("agente")

# abas de mês de 2026 (Abril26, Maio26, ... Dezembro26)
_ABA_MES = re.compile(r"26\s*$")

def _compareceu(status: str):
    s = _norm(status)
    if not s:
        return None
    if s.startswith(("vendido", "realizado", "finalizado", "entregue")):
        return True
    if s.startswith(("cancel", "faltou")) or "nao veio" in s or "nao compareceu" in s:
        return False
    return None  # agendado, negociação, ligação, etc.

_HEADER_MAP = {
    "cliente": "cliente", "vendedor": "vendedor", "data": "data",
    "horario": "hora", "status agendamento": "status", "status": "status",
    "veiculo": "veiculo", "canal de venda": "canal",
}


def _norm(s) -> str:
    s = str(s or "").strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def _parse_dt(dv, hv) -> str | None:
    d = None
    if isinstance(dv, datetime):
        d = dv
    elif isinstance(dv, date):
        d = datetime(dv.year, dv.month, dv.day)
    elif isinstance(dv, str) and dv.strip():
        for f in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
            try:
                d = datetime.strptime(dv.strip(), f)
                break
            except ValueError:
                continue
    if d is None:
        return None
    if isinstance(hv, time):
        d = d.replace(hour=hv.hour, minute=hv.minute)
    elif isinstance(hv, str):
        try:
            t = datetime.strptime(hv.strip(), "%H:%M")
            d = d.replace(hour=t.hour, minute=t.minute)
        except ValueError:
            pass
    return d.isoformat()


def _colmap(linha) -> dict:
    cols: dict[str, int] = {}
    for i, cel in enumerate(linha):
        chave = _HEADER_MAP.get(_norm(cel))
        if chave and chave not in cols:
            cols[chave] = i
    return cols


def _ref(aba: str, cliente: str, data_str: str, vendedor: str) -> str:
    base = f"{aba}|{cliente}|{data_str}|{vendedor}".strip().lower()
    return "plan_" + hashlib.md5(base.encode()).hexdigest()[:16]


def sincronizar() -> dict:
    url = f"https://docs.google.com/spreadsheets/d/{settings.planilha_sheet_id}/export?format=xlsx"
    r = httpx.get(url, timeout=90, verify=settings.verify_ssl, follow_redirects=True)
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)

    registros: list[dict] = []
    vistos: set[str] = set()
    abas = [n for n in wb.sheetnames if _ABA_MES.search(n)]

    # comparecimento já registrado (ex: vindo do grupo) p/ não ser apagado quando a planilha ainda diz "Agendado"
    compareceu_atual = {x["ref_externa"]: x.get("compareceu")
                        for x in db.select("agendamentos", {"select": "ref_externa,compareceu", "origem": "eq.planilha"})
                        if x.get("ref_externa")}

    # cache de vendedores (resolução em memória, sem ir ao banco por linha)
    vendedores = [v for v in db.select("vendedores", {"select": "id,nome,apelidos,funcao", "ativo": "eq.true"})
                  if v["funcao"] == "vendedor"]

    def resolver(nome: str):
        n = _norm(nome)
        if not n:
            return None
        for v in vendedores:
            if _norm(v["nome"]) == n or n in [_norm(a) for a in (v.get("apelidos") or [])]:
                return v
        for v in vendedores:
            if n in _norm(v["nome"]) or _norm(v["nome"]) in n:
                return v
        return None

    for aba in abas:
        ws = wb[aba]
        cols = None
        vazios = 0
        for linha in ws.iter_rows(values_only=True):
            if cols is None:
                if any(_norm(c) == "cliente" for c in linha):
                    cols = _colmap(linha)
                continue
            if "cliente" not in cols or "data" not in cols:
                break

            def val(k):
                i = cols.get(k)
                return linha[i] if i is not None and i < len(linha) else None

            cliente = str(val("cliente") or "").strip()
            data_raw = val("data")
            if not cliente or data_raw in (None, ""):
                vazios += 1
                if vazios > 30:  # fim dos dados da aba
                    break
                continue
            vazios = 0

            vendedor_nome = str(val("vendedor") or "").strip()
            status = str(val("status") or "").strip()
            veiculo = str(val("veiculo") or "").strip()
            canal = str(val("canal") or "").strip()
            data_iso = _parse_dt(data_raw, val("hora"))
            data_chave = data_iso[:10] if data_iso else str(data_raw)

            ref = _ref(aba, cliente, data_chave, vendedor_nome)
            if ref in vistos:  # evita ref duplicado no mesmo lote
                continue
            vistos.add(ref)
            vend = resolver(vendedor_nome)
            comp = _compareceu(status)
            if comp is None:  # planilha sem definição → preserva o que já foi marcado (ex: pelo grupo)
                comp = compareceu_atual.get(ref)
            registros.append({
                "cliente_nome": cliente,
                "vendedor_id": vend["id"] if vend else None,
                "data_agendada": data_iso,
                "compareceu": comp,
                "resultado": status or None,
                "origem": "planilha",
                "ref_externa": ref,
                "observacoes": " · ".join([p for p in (aba, status, veiculo, canal) if p]),
            })

    db.upsert("agendamentos", registros, "ref_externa")

    # espelha exclusões/edições: remove da base os da planilha que não estão mais nela
    existentes = db.select("agendamentos", {"select": "ref_externa", "origem": "eq.planilha"})
    remover = [e["ref_externa"] for e in existentes if e["ref_externa"] and e["ref_externa"] not in vistos]
    for i in range(0, len(remover), 100):
        lote = remover[i:i + 100]
        db.delete("agendamentos", {"origem": "eq.planilha", "ref_externa": f"in.({','.join(lote)})"})

    return {"abas": abas, "sincronizados": len(registros), "removidos": len(remover)}
